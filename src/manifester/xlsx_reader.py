import logging

from openpyxl import load_workbook
from manifester.xlsx_row import XLSXRow


def read_excel(xlsx_file: str) -> list[XLSXRow]:
    """
    Read Excel file into source records

    :param xlsx_file: full path to Excel file
    :return: list[XLSXRow] a list of source records
    """
    workbook = load_workbook(filename=xlsx_file)
    manifest_sheet = workbook["Manifest"]

    records = []
    for i in range(2, manifest_sheet.max_row):
        record = XLSXRow(manifest_sheet[i])
        records.append(record)
        logging.info(f'Extracted {record.identifier} ::: {record.title}')

    return records